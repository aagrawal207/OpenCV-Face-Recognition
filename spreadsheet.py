from openpyxl import Workbook, load_workbook
import time
import os
import sqlite3

#database connection
conn = sqlite3.connect('/home/malhar/Desktop/OpenCV-Face-Recognition/Face-DataBase')
c = conn.cursor()

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if(os.path.exists("reports.xlsx")):
    wb = load_workbook(filename = "reports.xlsx")
else:
    wb = Workbook()
    dest_filename = 'reports.xlsx'
    c.execute("SELECT * FROM Students ORDER BY Roll ASC")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "Cse15"
    ws1.append(('Roll Number', 'Name', '', currentDate))
    ws1.append(('', '', '', ''))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    #saving the file
    wb.save(filename = dest_filename)
